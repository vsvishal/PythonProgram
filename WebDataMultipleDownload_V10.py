import tkinter as tk
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from selenium import webdriver
import openpyxl
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
import os

# --------------------------------Tyco-----------------------------------------------------

class MutlipleWebDataDownload:


    def MainFunction(self):

        HEIGHT = 400
        WIDTH = 500

        root = tk.Tk()

        def btnClick():
            loopPartNum()


        root.title("Website Web Scrapping")

        canvas = tk.Canvas(root, height=HEIGHT, width=WIDTH)
        canvas.pack()

        btn = tk.Button(root, text="Download data", bg='#a2c49e', command=btnClick, font=1)
        # btn = tk.Button(root, text="Download data", bg='#a2c49e', font=1)
        btn.place(relx=0.3, rely=0.2, relheight=0.1, relwidth=0.3)

        lower_frame = tk.Frame(root, bg='#ffde00', bd=5)
        lower_frame.place(relx=0.5, rely=0.45, relwidth=0.85, relheight=0.30, anchor='n')

        labelLowerFrame = tk.Label(lower_frame, bg='white', font="Helvetica 10")
        labelLowerFrame.place(relwidth=1, height=110)

    #---------------------------------------------------------------------

        def loopPartNum():

            tyco_dict = {}
            molex_dict = {}
            aptiv_dict = {}
            deutsch_dict = {}
            yazaki_dict = {}
            path = r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Multiple_Data_Download.xlsx"
            pd_excel = pd.read_excel(path, sheet_name='Template')
            pd_count = pd_excel.count()
            pd_row_num = pd_count.iloc[0]

            #print(pd_row_num, pd_count)
            labelLowerFrame['text'] = 'Running completed.'

            for i in range(pd_row_num):
                #part_no = pd_excel.iloc[i]['Part Number']
                part_no = pd_excel.loc[i, 'Part Number']
                website_name = pd_excel.loc[i, 'Website']
                #print(part_no, i)
                if website_name == 'Tyco':
                    tyco_dict[i+2] = part_no
                    #TEAutomateTask(part_no, i+2, website_name)
                elif website_name == 'Molex':
                    molex_dict[i+2] = part_no
                    #MolexAutoTask(part_no, i+2, website_name)
                elif website_name == 'Aptiv(Delphi)':
                    aptiv_dict[i+2] = part_no
                    #AptivAutomateTask(part_no, i+2, website_name)
                elif website_name == 'Deutsch':
                    deutsch_dict[i+2] = part_no
                    #Deutsch(part_no, i+2, website_name)
                elif website_name == 'Yazaki':
                    yazaki_dict[i+2] = part_no
                    #Yazaki(part_no, i+2, website_name)
                else:
                    labelLowerFrame['text'] = 'Wrong website selected'+'\n'+ 'or you forgot to enter website name'

            #Store dictonary empty status
            tyco_dict_status = str(not bool(tyco_dict))
            molex_dict_status = str(not bool(molex_dict))
            aptiv_dict_status = str(not bool(aptiv_dict))
            deutsch_dict_status = str(not bool(deutsch_dict))
            yazaki_dict_status = str(not bool(yazaki_dict))

            #Check website dictonary is empty or not and call its function accrodingly
            if tyco_dict_status == "False":
                TEAutomateTask(tyco_dict, 'Tyco')

            if molex_dict_status == "False":
                MolexAutoTask(molex_dict, 'Molex')

            if aptiv_dict_status == "False":
                AptivAutomateTask(aptiv_dict, 'Aptiv(Delphi)')

            if deutsch_dict_status == "False":
                Deutsch(deutsch_dict, 'Deutsch')

            if yazaki_dict_status == "False":
                 Yazaki(yazaki_dict, 'Yazaki')

            #labelLowerFrame['text'] = 'Running completed.'


        # Pasting Attributes value in excel
        def putAttibuteValuesInExcel(r_no, attrib_value, wb_path, sh_name, col_num, part_no, website_name, euRohs_data, statusValue):
            wbkName = wb_path
            wbk = openpyxl.load_workbook(wbkName, data_only=True)
            wrksht = wbk[sh_name]

            if website_name == 'Tyco' or website_name == 'Deutsch':
                attrib_data = attrib_value
            else:
                attrib_data = attrib_value.encode('ascii', 'ignore').decode('ascii')

            # insert part number
            if website_name == 'Aptiv(Delphi)' or website_name == 'Deutsch' or website_name == 'Yazaki':
                pass
            else:
                wrksht.cell(row=2, column=col_num).value = part_no

            # insert status Value
            if website_name == 'Tyco':
                wrksht.cell(row=3, column=col_num).value = statusValue
            else:
                pass

            # insert Eu RoHS data
            if website_name == 'Molex' or website_name == 'Tyco':
                wrksht.cell(row=4, column=col_num).value = euRohs_data
            else:
                pass

            wrksht.cell(row=r_no + 2, column=col_num).value = attrib_data

            wbk.save(wbkName)


        def putValuesInTemplateFile(drawMsg,stepMsg,attributeMsg, statusMsg, row_num):
             wbkName = r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Multiple_Data_Download.xlsx"
             wbk = openpyxl.load_workbook(wbkName)
             wrksht = wbk['Template']

             wrksht.cell(row=row_num, column=3).value = drawMsg
             wrksht.cell(row=row_num, column=4).value = stepMsg
             wrksht.cell(row=row_num, column=5).value = attributeMsg
             wrksht.cell(row=row_num, column=6).value = statusMsg

             wbk.save(wbkName)


        #def TEAutomateTask(part_num, row_num, website_name):
        def TEAutomateTask(tyco_dict , website_name):

            chromeOptions = Options()

            prefs = {
                    "download.default_directory": r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Tyco",
                "plugins.always_open_pdf_externally": True
            }
            chromeOptions.add_experimental_option("prefs", prefs)
            baseUrl = "https://www.te.com/global-en/home.html"
            driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=chromeOptions)

            driver.implicitly_wait(10)
            driver.maximize_window()
            driver.get(baseUrl)
            time.sleep(5)

            for part in tyco_dict.items():

                partNo = str(part[1]).strip()
                print('p.no', partNo)
                row_num = part[0]
                searchFieldID = "search-input"
                submitButtonXpath = "//input[@type='submit']"
                drawingXpath = "//a[contains(@href,'Drawing') and contains(@href,'pdf')]"
                continueGuestXpath = "//div[@id='modal-content']//span[@class='guest-click ng-binding' and @data-ng-click='guestClicked()']"
                documentsXpath = "//a[@id='pdp-documents-tab']//span[contains(text(),'Documents')]"
                tb_data = []
                # noThanksPopupXpath = "//span[contains(text(),'No Thanks')]"
                stepFile = []
                stepFileSubText = '_stp'
                wb_name = r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Tyco_Attribute_Data.xlsx"
                sht_name = 'Data'

                driver.implicitly_wait(10)
                driver.maximize_window()
                driver.get(baseUrl)
                time.sleep(5)

                searchFieldElement = driver.find_element_by_id(searchFieldID)
                submitButtonElement = driver.find_element_by_xpath(submitButtonXpath)

                try:
                    searchFieldElement.clear()
                    searchFieldElement.send_keys(partNo)
                except Exception as e:
                    print(e)
                    te_dict = {}
                    te_dict = {part[0]:part[1]}
                    TEAutomateTask(te_dict, website_name)

                try:
                    # Switch the control to the Alert window
                    obj = driver.switch_to.alert
                    # Dismiss the Alert using
                    obj.dismiss()
                except:
                    pass

                # Submit button
                try:
                    if driver.find_elements_by_xpath(submitButtonXpath):
                        #submitButtonElement.click()
                        driver.execute_script("arguments[0].click();", submitButtonElement)
                except:
                    pass

                try:
                    # Switch the control to the Alert window
                    obj = driver.switch_to.alert
                    # Dismiss the Alert using
                    obj.dismiss()
                except:
                    pass

                time.sleep(5)

                # Check for double element search option
                if driver.find_elements_by_xpath("//span[contains(text(),'Features')]"):
                    pass
                else:
                    try:
                        obj = driver.switch_to.alert
                        # Dismiss the Alert using
                        obj.dismiss()
                    except:
                        pass
                    time.sleep(5)
                    try:
                        link_part = driver.find_elements_by_tag_name('a')
                        for link in range(len(link_part)):
                            if driver.find_elements_by_xpath('//div[@id="te-body"]//div//section[3]//table/tbody//div//span/a')[link].text == partNo:
                                link_element = driver.find_elements_by_xpath('//div[@id="te-body"]//div//section[3]//table/tbody//div//span/a')[link]
                                driver.execute_script("arguments[0].click();", link_element)
                    except:
                        try:
                            link_part2 = driver.find_elements_by_xpath("//div[@id='te-body']//div[@class='table-inner-div']//span[@class='internal-number product-values']")
                            print("Link part 2 len : ", len(link_part2))
                            for lnk in range(len(link_part2)):
                                if driver.find_elements_by_xpath("//div[@id='te-body']//div[@class='table-inner-div']//span[@class='internal-number product-values']")[lnk].text == partNo:
                                    link_element2 = driver.find_elements_by_xpath("//div[@id='te-body']//div[@class='table-inner-div']//span[@class='internal-number product-values']/parent::p/preceding-sibling::a")[lnk]
                                    driver.execute_script("arguments[0].click();", link_element2)
                        except:
                            pass
                        time.sleep(5)

                        try:
                            link_part3 = driver.find_elements_by_xpath("//div[@id='te-page']//div[@class='table2-inner-div2']//td[@class='ng-scope']//a")
                            for lnk2 in range(len(link_part3)):
                                if driver.find_elements_by_xpath("//div[@id='te-page']//div[@class='table2-inner-div2']//td[@class='ng-scope']//a")[lnk2].text == partNo:
                                    link_element3 = driver.find_elements_by_xpath("//div[@id='te-page']//div[@class='table2-inner-div2']//td[@class='ng-scope']//a")[lnk2]
                                    driver.execute_script("arguments[0].click();", link_element3)
                        except:
                            pass

                time.sleep(7)

                # Check whether part number present or not by verfiying Documents tab
                if driver.find_elements_by_xpath(documentsXpath):

                    statusObsoleteValue = driver.find_element_by_xpath("//ul[@class='statusContainer']//a[@id='dyn']//span").text

                    if driver.find_element_by_xpath("//ul[@class='statusContainer']//a[@id='dyn']//span").text == "ACTIVE":

                        # Download Drawing File
                        try:
                            # Switch the control to the Alert window
                            obj = driver.switch_to.alert
                            # Dismiss the Alert using
                            obj.dismiss()
                        except:
                            pass

                        try:
                            # drawingFeildElement = driver.find_element_by_xpath(drawingXpath)
                            # drawingFeildElement.click()

                            drawingElement = driver.find_element_by_xpath(drawingXpath)
                            driver.execute_script("arguments[0].click();", drawingElement)

                            # element = driver.find_element_by_css('div[class*="loadingWhiteBox"]')
                            # webdriver.ActionChains(driver).move_to_element(element).click(element).perform()

                            if driver.find_elements(By.XPATH, drawingXpath):
                                drawMsg = "Yes"
                        except Exception as e:
                            print(e)
                            drawMsg = "No"

                        try:
                            # Switch the control to the Alert window
                            obj = driver.switch_to.alert
                            # Dismiss the Alert using
                            obj.dismiss()
                        except:
                            pass

                        try:
                            # continueGuestElement = driver.find_element(By.CSS_SELECTOR, continueGuestCSS)
                            # continueGuestElement.click()
                            continueGuestElement = driver.find_element_by_xpath(continueGuestXpath)
                            driver.execute_script("arguments[0].click();", continueGuestElement)
                        except:
                            pass

                        try:
                            # Switch the control to the Alert window
                            obj = driver.switch_to.alert
                            # Dismiss the Alert using
                            obj.dismiss()
                        except:
                            pass

                        time.sleep(8)

                        # Rename drawing file name
                        try:
                            elem = driver.find_element_by_xpath(drawingXpath)
                            source = elem.get_attribute('outerHTML')

                            source_split = source.split()
                            source_split_2 = (source_split[2]).split("=")

                            # Get drawing file name
                            link_draw = source_split_2[-1]
                            split_link = link_draw.split("%7F")
                            drawingDownloadFileName = split_link[-2]

                            # Get extension of drawing file
                            extension = os.path.splitext(drawingDownloadFileName)

                            # Store new file name
                            newFileName = str(partNo) + str(extension[-1])

                            # Change directory and rename downloaded file:
                            os.chdir(r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Tyco")
                            os.renames(drawingDownloadFileName, newFileName)
                        except Exception as e:
                            print(e)

                        # download Step File
                        try:
                            # Switch the control to the Alert window
                            obj = driver.switch_to.alert
                            # Dismiss the Alert using
                            obj.dismiss()
                        except:
                            pass

                        try:
                            for a in driver.find_elements_by_xpath('.//a'):
                                stepFile.append(a.get_attribute('href'))

                            res = []
                            for val in stepFile:
                                if val != None:
                                    res.append(val)

                            for i in res:
                                if stepFileSubText in i:
                                    stpZip = i
                            # time.sleep(5)
                            driver.get(stpZip)  # Getting Step File URL for Downloading file
                            if not stpZip == "":
                                stepFileMsg = "Yes"
                        except:
                            stepFileMsg = "No"

                        time.sleep(5)

                        #global euRohs_data
                        # Tab Eu RoHS data
                        try:
                            # click on Product Compliance Tab
                            #driver.find_element_by_css_selector('#pdp-compliance-tab').click()
                            productCompliancetElement = driver.find_element_by_css_selector('#pdp-compliance-tab')
                            driver.execute_script("arguments[0].click();", productCompliancetElement)

                            # Get EU RoHS text
                            euRoHS = driver.find_element_by_xpath(
                                "//div[@class='product-document-type overwritePadding']//div[@class='document-type-label-column  is-current-language']//h3").text

                            euRoHSValue = driver.find_element_by_xpath(
                                "//div[@class='product-document-type overwritePadding']//a[@data-tn-compliance-btn='rohs10']").text

                            euRoHS_data = euRoHS + " : " + euRoHSValue
                        except Exception as e:
                            print(e)

                        time.sleep(7)


                        # Click on Features Tab
                        try:
                            if driver.find_elements_by_xpath("//span[contains(text(),'Features')]"):
                                #driver.find_element_by_xpath("//span[contains(text(),'Features')]").click()
                                feature_tab = driver.find_element_by_xpath("//span[contains(text(),'Features')]")
                                driver.execute_script("arguments[0].click();", feature_tab)
                        except:
                            pass
                            # statusMsg = partNo + ' ' + 'is not present on Tyco Website'
                            # #partNotFoundMsg(msg)
                            # driver.close()
                            # driver.quit()

                        try:
                            # Switch the control to the Alert window
                            obj = driver.switch_to.alert
                            # Dismiss the Alert using
                            obj.dismiss()
                        except:
                            pass

                        time.sleep(5)

                        try:
                            # get status
                            statusValue = driver.find_element_by_xpath("//ul[@class='statusContainer']//a[@id='dyn']//span").text

                            table = driver.find_element_by_css_selector("div#pdp-features-tabpanel")
                            items = table.find_elements_by_tag_name("li")
                            tb_data.append([d.text for d in items])

                            # pd_data = pd.DataFrame(tb_data, columns=['Attribute', 'Values'])
                            pd_data = pd.DataFrame(tb_data)
                            pd_transpose = pd_data.transpose()

                            pd_transpose.columns = ['DataValues']

                            # Dropping null values to avoid errors
                            pd_transpose.dropna(inplace=True)

                            # new data frame with split value columns
                            pd_split_data = pd_transpose['DataValues'].str.split(':', n=1, expand=True)
                            pd_split_data.columns = ['Attribute', 'Values']

                            #get max colm num
                            path = wb_name
                            wbk = openpyxl.load_workbook(path)
                            wrksht = wbk[sht_name]
                            colm_num = (wrksht.max_column) + 1
                            wbk.close()

                            # extract Attribute Values
                            pd_excel = pd.read_excel(path, sheet_name=sht_name)
                            pd_count = pd_excel.count()
                            pd_row_num = pd_count.iloc[0]

                            print(pd_row_num)

                            for i in range(pd_row_num):
                                attrib_name = pd_excel.loc[i, 'Attribute Name']
                                try:
                                    attrib_value_row = pd_split_data.loc[pd_split_data['Attribute'] == attrib_name]
                                    attrib_ind = pd_split_data.index[pd_split_data['Attribute'] == attrib_name].tolist()
                                    #print('attrib_ind', attrib_ind)
                                    attrib_index = attrib_ind[0]  # get row index from DataFrame
                                    attrib_value = attrib_value_row.loc[attrib_index, 'Values']
                                    putAttibuteValuesInExcel(i, attrib_value, path, sht_name, colm_num, partNo, website_name,euRoHS_data, statusValue)

                                except Exception as e:
                                    #print(e)
                                    continue

                            if driver.find_elements_by_css_selector("div#pdp-features-tabpanel"):
                                attributeMsg = "Yes"

                        except:
                            attributeMsg = "No"

                        time.sleep(8)

                        statusMsg = 'Yes'
                        putValuesInTemplateFile(drawMsg, stepFileMsg, attributeMsg, statusMsg, row_num)

                    else:
                        statusMsg = statusObsoleteValue
                        putValuesInTemplateFile('No','No','No',statusMsg,row_num)

                else:
                    statusMsg = 'No'
                    putValuesInTemplateFile('No', 'No', 'No', statusMsg, row_num)

            #Close the browser
            driver.close()
            driver.quit()

        # -------------------------------------------------------------------------------------

        # ----------------------------Molex----------------------------------------------------

        def MolexAutoTask(molex_dict, website_name):

            chromeOptions = Options()
            prefs = {
                "download.default_directory": r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Molex",
                "plugins.always_open_pdf_externally": True
            }
            chromeOptions.add_experimental_option("prefs", prefs)

            baseUrl = "https://www.molex.com/molex/home"
            driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=chromeOptions)

            driver.maximize_window()
            driver.get(baseUrl)
            driver.implicitly_wait(10)
            time.sleep(6)

            cnt = 0
            for part in molex_dict.items():

                partNo = str(part[1]).strip()
                row_num = part[0]
                #print('Pno', partNo)
                partNoID = "query"
                partNoInsideId = "search-input"
                searchButtonClass = "input-group-btn"
                searchBtnXpath = "//button[@class='btn btn-primary btn-plain' and @type='submit' and text()='Search']"
                serachBtnInsideXpath = "//ul[@id='nav-left']//a[@href='#search-form']"
                drawingFileXpath = "//a[text()='Drawing (PDF)']"
                model3DXpath = "//a[text()='3D Model']"
                stepFileXpath = "//a[text()='3D CAD Models - STEP']"
                statusXpath = "//div[@id='part-overview']/div/div[1]//div[@class='status-wrap vam']/div/span"
                EuRoHSXpath = "//div[@class='section space-sm plain display-box js_scroll_to_section-3']//div[@class='main']//div//p[3]"
                tb_data = []
                mates_data = []
                euRohs_data = ''
                workbook_path = r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Molex_Attribute_Data.xlsx"
                sht_name = "Data"

                #Click on search icon present inside
                try:
                    if cnt > 0:
                        serachBtnInsideElement = driver.find_element_by_xpath(serachBtnInsideXpath)
                        serachBtnInsideElement.click()
                except:
                    print(e)
                    molex_dict = {}
                    molex_dict = {part[0]:part[1]}
                    MolexAutoTask(molex_dict, website_name)

                if cnt == 0:
                    partNoElement = driver.find_element(By.ID, partNoID)
                    partNoElement.clear()
                    partNoElement.send_keys(partNo)
                else:
                    partNoElement = driver.find_element(By.ID, partNoInsideId)
                    partNoElement.clear()
                    partNoElement.send_keys(partNo)


                if cnt == 0:
                    searchButtonElement = driver.find_element(By.CLASS_NAME, searchButtonClass)
                    searchButtonElement.click()
                else:
                    searchButtonElement2 = driver.find_element(By.XPATH, searchBtnXpath)
                    searchButtonElement2.click()

                time.sleep(6)

                #increase count
                cnt += 1

                try:
                    if driver.find_elements_by_xpath("//div[@id='part-overview']//strong[text()='Status:']"):

                        #Click on expand all button
                        try:
                            driver.execute_script("document.getElementsByClassName('btn btn-primary js_expand_all')[0].click()")
                            time.sleep(4)
                        except:
                            pass

                        # Get attribute data
                        try:
                            table = driver.find_element_by_xpath(
                                "//div[@class='section space-sm plain display-box js_scroll_to_section-5']")
                            for row in table.find_elements_by_css_selector('tr'):
                                tb_data.append([d.text for d in row.find_elements_by_css_selector('td')])

                            pd_data = pd.DataFrame(tb_data, columns=['Attribute', 'Values'])

                            statusValue = driver.find_element_by_xpath(statusXpath).text
                            if driver.find_element_by_xpath(statusXpath).text == 'Active' or driver.find_element_by_xpath(statusXpath).text == "Active - Custom":

                                # get max colm num
                                path = workbook_path
                                wbk = openpyxl.load_workbook(path)
                                wrksht = wbk[sht_name]
                                col_num = wrksht.max_column+1
                                wbk.close()

                                #path = workbook_path
                                pd_excel = pd.read_excel(path, sheet_name=sht_name)
                                pd_count = pd_excel.count()
                                pd_row_num = pd_count.iloc[0]

                                # Eu RoHS data
                                try:
                                    euRohs = driver.find_element_by_xpath(EuRoHSXpath).text
                                    euRohs_value = euRohs.split(":")
                                    euRohs_value[1].strip(" ")
                                    euRohs_data = euRohs_value[1]
                                except Exception as e:
                                    print(e)

                                #Match Pandas attribute data with attribute file value one by one and paste in Attribute file
                                for i in range(pd_row_num):
                                    attrib_name = pd_excel.loc[i, 'Attribute Name']
                                    try:
                                        attrib_value_row = pd_data.loc[pd_data['Attribute'] == attrib_name]
                                        attrib_ind = pd_data.index[pd_data['Attribute'] == attrib_name].tolist()
                                        attrib_index = attrib_ind[0]
                                        attrib_value = attrib_value_row.loc[attrib_index, 'Values']
                                        putAttibuteValuesInExcel(i, attrib_value, path, sht_name, col_num, partNo, website_name, euRohs_data, 'Status')
                                    except Exception as e:
                                        #print(e)
                                        continue

                                    try:
                                        mates_table = driver.find_element_by_css_selector(
                                            '#inner-content > div.section.space-sm.plain.display-box.js_scroll_to_section-2')
                                        mates_p = mates_table.find_elements_by_tag_name("p")
                                        mates_data.append([d.text for d in mates_p])

                                        pd_mates = pd.DataFrame(mates_data)

                                        pd_transpose = pd_mates.transpose()
                                        pd_transpose.columns = ['Mating Values']

                                        writer = pd.ExcelWriter(path, engine='openpyxl')
                                        book = load_workbook(path)
                                        writer.book = book
                                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                                        pd_transpose.to_excel(writer, sheet_name=sht_name, header=None, index=False, startrow=pd_row_num + 1,
                                                              startcol=col_num-1, encoding='utf-8')
                                        writer.save()

                                    except:
                                        pass
                                        #print(e)

                                if driver.find_elements_by_xpath("//div[@class='section space-sm plain display-box js_scroll_to_section-5']"):
                                    attributeMsg = "Yes"
                                else:
                                    attributeMsg = "No"

                                time.sleep(8)


                                # Drawing file download
                                try:
                                    drawingElement = driver.find_element(By.XPATH, drawingFileXpath)
                                    drawingElement.click()

                                    if driver.find_elements(By.XPATH, drawingFileXpath):
                                        drawMsg = "Yes"
                                except:
                                    drawMsg = "No"

                                time.sleep(4)

                                # Step File Download
                                try:
                                    model3DElement = driver.find_element(By.XPATH, model3DXpath)
                                    model3DElement.click()
                                    driver.find_element_by_name("AGREE").click()
                                    stepFileElement = driver.find_element(By.XPATH, stepFileXpath)
                                    stepFileElement.click()

                                    if driver.find_elements(By.XPATH, stepFileXpath):
                                        stepFileMsg = "Yes"
                                except:
                                    stepFileMsg = "No"

                                time.sleep(8)
                                statusMsg = 'Yes'
                                putValuesInTemplateFile(drawMsg, stepFileMsg, attributeMsg, statusMsg, row_num)

                            # If Status Obsolete than close
                            else:
                                statusMsg = statusValue
                                putValuesInTemplateFile('No', 'No', 'No', statusMsg, row_num)
                                time.sleep(4)

                        except Exception as e:
                            pass
                            #print(e)
                            #attributeMsg = "No"
                    else:
                        statusMsg = 'No'
                        print("hello VS")
                        putValuesInTemplateFile('No', 'No','No', statusMsg, row_num)

                except Exception as e:
                    print(e, "Something went wrong")

            driver.close()
            driver.quit()

        # -------------------------------------------------------------------------------------
        # ---------------------------Aptiv(Delphi)---------------------------------------------
        def AptivAutomateTask(aptiv_dict, website_name):
            chromeOptions = Options()

            prefs = {
                 "download.default_directory": r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Aptiv(Delphi)",
                 "plugins.always_open_pdf_externally": True
            }

            chromeOptions.add_experimental_option("prefs", prefs)
            baseUrl = "https://ecat.aptiv.com"
            driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=chromeOptions)

            driver.implicitly_wait(10)
            driver.maximize_window()
            driver.get(baseUrl)              # Got to website
            time.sleep(5)

            for part in aptiv_dict.items():

                partNo = str(part[1]).strip()
                row_num = part[0]
                print(partNo)
                searchFieldID = "searchUserInput"
                cookiesCssSelector = "button#tracking-consent-dialog-reject"
                documentsXpath = "//a[@title='Product Documentation' and @ng-click='ViewProductDocuments()']"
                stpFileXpath = "//a[contains(text(),'_STP')]"
                drawingXpath = "//a[contains(text(),'2D_PDF')]"
                drawingTiffCssSelector = "a[href*='tif']"
                noResultXpath = "//div[@class='col-lg-3 ng-scope']//strong[@class='ng-binding']"
                partImageClickXpath = "//div[@id='Contentplaceholder1_TE1068189018_Col00']//span//img"
                relatedProduct = "//a[@title='Related Products' and @ng-click='ViewRelatedProducts()']"
                tb_data = []
                tb_related_data = []
                wrk_book_path = r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Aptiv(Delphi)_Attribute_Data.xlsx"
                sht_name = "Data"

                try:
                    partNoElement = driver.find_element(By.ID, searchFieldID)
                    partNoElement.clear()
                    partNoElement.send_keys(partNo)
                except Exception as e:
                    print(e)
                    aptiv_dict = {}
                    aptiv_dict = {part[0]: part[1]}
                    AptivAutomateTask(aptiv_dict, website_name)

                time.sleep(3)

                try:
                    cookiesElement = driver.find_element(By.CSS_SELECTOR, cookiesCssSelector)
                    cookiesElement.click()
                except:
                    pass

                # click on search button
                try:
                    driver.execute_script("document.getElementsByClassName('fa fa-search')[0].click()")
                except:
                    pass

                time.sleep(5)

                # Checks whether Part is found or not
                try:
                    zeroResultElement = driver.find_element_by_xpath(noResultXpath).text
                    if zeroResultElement == partNo:

                        # click on image
                        driver.find_element_by_xpath(partImageClickXpath).click()

                        time.sleep(5)

                        # switch to new tab
                        driver.switch_to.window(driver.window_handles[1])

                        time.sleep(5)

                        # Create csv file and paste attributes in it
                        try:
                            table = driver.find_element_by_css_selector("table[class='ProductDocumentsTable']")
                            for row in table.find_elements_by_css_selector('tr'):
                                tb_data.append([d.text for d in row.find_elements_by_css_selector('td')])

                            pd_data = pd.DataFrame(tb_data, columns=['Attribute', 'Values'])
                            pd_data.apply(lambda x: x.str.strip())


                            # get max colm num
                            path = wrk_book_path
                            wbk = openpyxl.load_workbook(path)
                            wrksht = wbk[sht_name]
                            col_num = (wrksht.max_column) + 1
                            wbk.close()

                            #path = wrk_book_path
                            pd_excel = pd.read_excel(path, sheet_name=sht_name)
                            pd_count = pd_excel.count()
                            pd_row_num = pd_count.iloc[0]

                            print(pd_row_num)

                            for i in range(pd_row_num):
                                attrib_name = pd_excel.loc[i, 'Attribute Name']
                                try:
                                    attrib_value_row = pd_data.loc[pd_data['Attribute'] == attrib_name]
                                    attrib_ind = pd_data.index[pd_data['Attribute'] == attrib_name].tolist()
                                    attrib_index = attrib_ind[0]
                                    attrib_value = attrib_value_row.loc[attrib_index, 'Values']
                                    putAttibuteValuesInExcel(i, attrib_value, path, sht_name, col_num, partNo, website_name,'No','Status')
                                except Exception as e:
                                    #print(e)
                                    continue

                            time.sleep(7)

                            try:
                                relatedProductElement = driver.find_element_by_xpath(relatedProduct)
                                driver.execute_script("arguments[0].click();", relatedProductElement)
                                #relatedProductElement.click()

                                # related product element
                                related_body = driver.find_element_by_xpath("//div[@class='ProductDetailsBody']")
                                for row in related_body.find_elements_by_tag_name('tr'):
                                    tb_related_data.append([d.text for d in row.find_elements_by_tag_name('td')])

                                pd_related = pd.DataFrame(tb_related_data, columns=['p', 'a', 'n', 'd'])
                                pd_related.to_csv(r'C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Aptive.csv', index=False)
                                pd_csv = pd.read_csv(r'C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Aptive.csv')
                                pd_csv.dropna(how='all', inplace=True)
                                #print(pd_csv)

                                writer = pd.ExcelWriter(wrk_book_path, engine='openpyxl')
                                book = load_workbook(wrk_book_path)
                                writer.book = book
                                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                                pd_csv.to_excel(writer, sheet_name=sht_name, header=None, index=False, startrow=pd_row_num,
                                                startcol=col_num - 1, encoding='utf-8')
                                writer.save()
                            except Exception as e:
                                print(e)

                            try:
                                os.remove(r'C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Aptive.csv')
                            except Exception as e:
                                print(e)


                            if driver.find_elements_by_css_selector("table[class='ProductDocumentsTable']"):
                                attributeMsg = "Yes"


                        except Exception as e:
                            #print(e)
                            attributeMsg = "No"

                        time.sleep(5)


                        # click on Documents tab
                        try:
                            documentElement = driver.find_element_by_xpath(documentsXpath)
                            driver.execute_script("arguments[0].click();", documentElement)
                        except Exception as e:
                            print(e)

                        # Step File download
                        try:
                            stpFileElement = driver.find_element_by_xpath(stpFileXpath)
                            #stpFileElement.click()
                            driver.execute_script("arguments[0].click();", stpFileElement)

                            if driver.find_elements_by_xpath(stpFileXpath):
                                stepFileMsg = "Yes"
                        except:
                            stepFileMsg = "No"

                        time.sleep(5)


                        # Drawing File Download
                        try:
                            if driver.find_elements_by_xpath(drawingXpath):
                                #driver.find_element_by_xpath(drawingXpath).click()
                                drawingPdf = driver.find_element_by_xpath(drawingXpath)
                                driver.execute_script("arguments[0].click();", drawingPdf)
                                drawMsg = "Yes"
                            else:
                                if driver.find_elements(By.CSS_SELECTOR, drawingTiffCssSelector)[0]:
                                    drawingElementTiff = driver.find_elements(By.CSS_SELECTOR, drawingTiffCssSelector)[0]
                                    driver.execute_script("arguments[0].click();", drawingElementTiff)
                                    drawMsg = "Yes"
                                elif driver.find_elements(By.CSS_SELECTOR, drawingTiffCssSelector)[1]:
                                    drawingElementTiff2 = driver.find_elements(By.CSS_SELECTOR, drawingTiffCssSelector)[1]
                                    driver.execute_script("arguments[0].click();", drawingElementTiff2)
                                    drawMsg = "Yes"
                                else:
                                    drawMsg = "No"
                            #drawMsg = "Yes"

                        except:
                            drawMsg = "No"

                        time.sleep(14)
                        statusMsg = 'Yes'
                        putValuesInTemplateFile(drawMsg,stepFileMsg, attributeMsg, statusMsg, row_num)

                    else:
                        #statusMsg = 'Part ' + partNo + ' ' + 'is not present on Molex Website'
                        statusMsg = 'No'
                        putValuesInTemplateFile('No','No','No',statusMsg,row_num)

                except Exception as e:
                    statusMsg = 'No'
                    putValuesInTemplateFile('No', 'No', 'No', statusMsg, row_num)
                    print(e, "Something went wrong")

                #Close the current tab in browser
                driver.close()

                #Switch to main browser tab
                try:
                    driver.switch_to.window(driver.window_handles[0])
                except Exception as e:
                    print(e)

            #Close the website
            driver.close()
            driver.quit()

        #-------------------------------Deutsch-------------------------------------------------------

        def Deutsch(deutsch_dict, website_name):

            chromeOptions = Options()
            prefs = {
                "download.default_directory": r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Deutsch",
                "plugins.always_open_pdf_externally": True
            }
            chromeOptions.add_experimental_option("prefs", prefs)
            driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=chromeOptions)

            #baseUrl = "https://laddinc.com/contact/"
            driver.implicitly_wait(10)
            driver.maximize_window()
            #driver.get(baseUrl)
            #time.sleep(8)

            for part in deutsch_dict.items():

                partNo = str(part[1]).strip()
                row_num = part[0]

                baseUrl = "https://laddinc.com/product/?sku=" + partNo
                driver.get(baseUrl)
                time.sleep(8)

                #searchBoxCss = "#search-form > form > input"
                #searchBtnCss = "#search-form > form > button > i"
                tb_data = []
                ModelsDrawingTabCss = "#drawings-tab"
                drawingPdfXpath = "//div[@id='drawings']//ul[@class='product-drawings']//a[contains(@href, 'pdf')]"
                drawingTifXpath = "//div[@id='drawings']//ul[@class='product-drawings']//a[contains(@href, 'tif')]"
                cadDrawingXpath = "//div[@id='drawings']//a[@data-target='#drawingModal' and contains(text(),'CAD Drawings')]"
                agreeBtnCss = "#accept-drawings"
                stpFileXpath = "//div[@id='drawing-links']//li[@class='icon step']//a[contains(@href, '3d_stp')]"
                technicalDetailsTabCss = "#details-tab"
                wb_name = r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Deutsch_Attribute_Data.xlsx"
                sht_name = "Data"

                # Check whether part is present on website
                if driver.find_elements_by_css_selector(technicalDetailsTabCss):

                    # Extract Attribute Data
                    try:
                        #global attributeMsg
                        if driver.find_elements_by_css_selector("#details > ul"):
                            attributeMsg = "Yes"

                        table = driver.find_element_by_css_selector("#details > ul")
                        items = table.find_elements_by_tag_name("li")
                        tb_data.append([d.text for d in items])

                        # pd_data = pd.DataFrame(tb_data, columns=['Attribute', 'Values'])
                        pd_data = pd.DataFrame(tb_data)
                        pd_transpose = pd_data.transpose()

                        pd_transpose.columns = ['DataValues']

                        # Dropping null values to avoid errors
                        pd_transpose.dropna(inplace=True)

                        # new data frame with split value columns
                        pd_split_data = pd_transpose['DataValues'].str.split(':', n=1, expand=True)
                        pd_split_data.columns = ['Attribute', 'Values']

                        # get max colm num
                        path = wb_name
                        wbk = openpyxl.load_workbook(path)
                        wrksht = wbk[sht_name]
                        colm_num = wrksht.max_column + 1
                        wbk.close()

                        # extract Attribute Values
                        pd_excel = pd.read_excel(path, sheet_name=sht_name)
                        pd_count = pd_excel.count()
                        pd_row_num = pd_count.iloc[0]

                        print(pd_row_num)

                        for i in range(pd_row_num):
                            attrib_name = pd_excel.loc[i, 'Attribute Name']
                            try:
                                attrib_value_row = pd_split_data.loc[pd_split_data['Attribute'] == attrib_name]
                                attrib_ind = pd_split_data.index[pd_split_data['Attribute'] == attrib_name].tolist()
                                # print('attrib_ind', attrib_ind)
                                attrib_index = attrib_ind[0]  # get row index from DataFrame
                                attrib_value = attrib_value_row.loc[attrib_index, 'Values']
                                putAttibuteValuesInExcel(i, attrib_value, path, sht_name, colm_num, partNo, website_name,
                                                         'No_euRoHS_data', 'No_statusValue')

                            except Exception as e:
                                # print(e)
                                continue

                    except:
                        attributeMsg = "No"

                    time.sleep(7)

                    # click on Models & Drawing Tab
                    try:
                        ModelsDrawingElement = driver.find_element_by_css_selector(ModelsDrawingTabCss)
                        driver.execute_script("arguments[0].click();", ModelsDrawingElement)
                    except Exception as e:
                        print(e)

                    # drawing file download
                    try:
                        #global drawMsg
                        if driver.find_element_by_xpath(drawingPdfXpath):
                            drawingPdfElement = driver.find_element_by_xpath(drawingPdfXpath)
                            driver.execute_script("arguments[0].click();", drawingPdfElement)
                            drawMsg = "Yes"
                        elif driver.find_element_by_xpath(drawingTifXpath):
                            drawingTifElement = driver.find_element_by_xpath(drawingTifXpath)
                            driver.execute_script("arguments[0].click();", drawingTifElement)
                            drawMsg = "Yes"
                        else:
                            drawMsg = "No"
                    except Exception as e:
                        print(e)
                    time.sleep(6)

                    # stp file download
                    try:
                        #global stepFileMsg
                        # Click on Cad Drawings
                        cadDrawingsElement = driver.find_element_by_xpath(cadDrawingXpath)
                        driver.execute_script("arguments[0].click();", cadDrawingsElement)

                        # Click on I Agree button
                        agreeBtnElement = driver.find_element_by_css_selector(agreeBtnCss)
                        driver.execute_script("arguments[0].click()", agreeBtnElement)

                        # Click on stp file
                        stpFileElement = driver.find_element_by_xpath(stpFileXpath)
                        driver.execute_script("arguments[0].click()", stpFileElement)

                        if driver.find_element_by_xpath(stpFileXpath):
                            stepFileMsg = "Yes"

                    except Exception as e:
                        stepFileMsg = "No"
                        print(e)

                    time.sleep(10)
                    statusMsg = 'Yes'
                    putValuesInTemplateFile(drawMsg, stepFileMsg, attributeMsg, statusMsg, row_num)

                else:
                    statusMsg = 'No'
                    putValuesInTemplateFile('No', 'No', 'No', statusMsg, row_num)

            #Close the brower
            driver.close()
            driver.quit()

        #--------------------------------------Yazaki-------------------------------------------------------------------

        def Yazaki(yazaki_dict, website_name):

            chromeOptions = Options()
            prefs = {
                "download.default_directory": r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Yazaki",
                "plugins.always_open_pdf_externally": True
            }
            chromeOptions.add_experimental_option("prefs", prefs)
            driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=chromeOptions)

            driver.implicitly_wait(10)
            driver.maximize_window()
            #driver.get(baseUrl)
            #time.sleep(6)

            for part in yazaki_dict.items():

                p_no = str(part[1])
                row_num = part[0]
                partNo = p_no.replace('-', '').replace(' ', '')
                tb_data = []
                drawingCss = "#cph1_lnkQuickDrawing"
                wb_path = r"C:\Users\VS\Documents\JD_Files\Wed Data Download\Attribute Files\Yazaki_Attribute_Data.xlsx"
                sht_name = "Data"
                baseUrl = "https://apps.us.yazaki.com/componentcatalog/Component_Detail.aspx?SearchItemPN=" + partNo
                driver.get(baseUrl)
                time.sleep(8)

                # Check whether part no preasent
                if driver.find_element_by_xpath("//div[@class='rightCol']"):

                    # Get Attribute Data
                    try:
                        table = driver.find_element_by_xpath("//div[@class='rightCol']")
                        items = table.find_elements_by_tag_name('tr')
                        tb_data.append([d.text for d in items])

                        # pd_data = pd.DataFrame(tb_data, columns=['Attribute', 'Values'])
                        pd_data = pd.DataFrame(tb_data)

                        pd_transpose = pd_data.transpose()
                        pd_transpose.columns = ['DataValues']

                        # Dropping null rows value to avoid errors
                        pd_transpose.dropna(axis=0, how='all', thresh=None, subset=None, inplace=True)

                        pd_split_data = pd_transpose['DataValues'].str.split(':', n=1, expand=True)
                        pd_split_data.columns = ['Attribute', 'Values']

                        # get max colm num
                        path = wb_path
                        wbk = openpyxl.load_workbook(path)
                        wrksht = wbk[sht_name]
                        colm_num = wrksht.max_column + 1
                        wbk.close()

                        # extract Attribute Values
                        pd_excel = pd.read_excel(path, sheet_name=sht_name)
                        pd_count = pd_excel.count()
                        pd_row_num = pd_count.iloc[0]

                        print(pd_row_num)

                        for i in range(pd_row_num):
                            attrib_name = pd_excel.loc[i, 'Attribute Name']
                            try:
                                attrib_value_row = pd_split_data.loc[pd_split_data['Attribute'] == attrib_name]
                                attrib_ind = pd_split_data.index[pd_split_data['Attribute'] == attrib_name].tolist()
                                attrib_index = attrib_ind[0]  # get row index from DataFrame
                                attrib_value = attrib_value_row.loc[attrib_index, 'Values']
                                putAttibuteValuesInExcel(i, attrib_value, path, sht_name, colm_num, partNo, website_name,
                                                         'No_euRoHS_data', 'No_statusValue')

                            except Exception as e:
                                # print(e)
                                continue
                        attributeMsg = 'Yes'

                    except:
                        attributeMsg = "No"

                    time.sleep(7)

                    # Drawing file download
                    try:
                        drawingElement = driver.find_element_by_css_selector(drawingCss)
                        driver.execute_script("arguments[0].click();", drawingElement)
                        drawMsg = "Yes"
                    except Exception as e:
                        drawMsg = "No"
                        print(e)

                    time.sleep(8)
                    statusMsg = 'Yes'
                    putValuesInTemplateFile(drawMsg, 'No', attributeMsg, statusMsg, row_num)
                    #driver.close()
                    #driver.quit()

                else:
                    statusMsg = 'No'
                    putValuesInTemplateFile('No', 'No', 'No', statusMsg, row_num)
                    #driver.close()
                    #driver.quit()

            #Close the browser
            driver.close()
            driver.quit()

        #---------------------------------------------------------------------------------------------------------------

        root.mainloop()

guiObj = MutlipleWebDataDownload()
guiObj.MainFunction()

#--------------------------------------------------The End--------------------------------------------------------------
